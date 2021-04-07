# """
# ==========
# Author:TT
# Time:2021/2/7  11:42 上午
# Project: API_ningmeng_dome
# Company:自动化测试
# ==========
#
# openpyxl
# 准备测试数据
# 平时操作excel的流程 （三个对象）
# 1；工作薄-->Workbokk
# 2：表单 --> Sheet
# 3：单元格 ---> Cell
#
#
# 1:准备测试数据
# 2：load_workbook，去打开测试数据文件 生成WorkBook对象
# 3，根据表单名称选择表单：wb['login']
# 4:在表单中 获取单元格数据：
#     4。1 单元格对象 ：sh.cell(row,cloum) 下标从1开始
#     4.2  .value 获取单元格的值
#     4.3
# 5：得到当前表单当中总行数和总列数
#     sh.max_row   #总行数
#     sh.max_column #总列数
# 6：修改数据 sh.cell(row,cloum).value = 新的值
#
# 7： 保存数据（保存整个工作薄）
#     WorkBook对象（wb）.save(文件路径）
#
# """
# import os
# file_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)),"Excel.xlsx")
# from openpyxl import load_workbook
# print(file_dir)
# #加载excel数据文件
# wb=load_workbook(file_dir)
#
# #2，根据表单名称选择表单 wb['login']
# sh=wb['login']
# print(sh)
#
# #3：获取表单数据
# cel = sh.cell(2,2).value
# print(cel)
#
# # sh.max_row   #总行数
# # sh.max_column #总列数
# print(sh.max_row)
# print(sh.max_column)
# print("===================")
#
# # 修改数据 sh.cell(row,cloum).value = 新的值
# sh.cell(2,2).value="1wergvcsxsc"
# print(sh.cell(2,2).value)


#按行读取数据
#     sh.rows = 所有行的数据。list(sh.rows)返回的是一个列表，列表当中返回的是一个成员，每一行的数据元组
# import os
# file_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)),"Excel.xlsx")
# from openpyxl import load_workbook
# print(file_dir)
# #加载excel数据文件
# wb=load_workbook(file_dir)
#
# #2，根据表单名称选择表单 wb['login']
# sh=wb['login']
# all_data = [] #获取excel表格当中的所有测试数据
# # #第一步：拿到字典的key值
# # print(list(sh.rows)[0])
# title = []
# for item in list(sh.rows)[0]:
#     title.append(item.value)
# print(title)
# #
# # #第二步；把key和value组合到一起，形成一个字典，放到列表当中
# # # # print(list(sh.rows)) #每一行都是元组，元组里放的是每一行的单元格
# # data_list=[]
# # for item in list(sh.rows)[1:]:
# #     # (<Cell 'login'.A1>, <Cell 'login'.B1>, <Cell 'login'.C1>)v
# #     # print(item)
# #     value_dict={}#每一行是一个字典
# #     # print(item)
# #     for index in range(len(item)): #获取每一行的单元格数据
# #         # print(index,item[index],item[index].value)
# #         value_dict[title[index]]=item[index].value
# #     # print(value_dict)
# #     data_list.append(value_dict)#将每一行测试数据追加到列表中
# # print(data_list)
#
# for item in list(sh.rows)[1:]: #遍历数据行
#     value = []
#     for val in item: #获取每一行的数据
#         value.append(val.value)
#     res = dict(zip(title,value))#title 和每一行数据打包成为字典
#     res['check']=eval(res['check']) #将check的字符串转换为字典对象
#     all_data.append(res) #追加数据
# print(all_data)
import json
import os
from common.Case_file import conf
from openpyxl import load_workbook
class HandlerExcel:
    """
    excel类，你的需求是实现什么
    1：读取表头
    2：读取数据 - 读取表头以外的数据，-返回值。：列表 成员是每一行数据

    初始化工作？ 加载一个Excel ，打开一个表单
    """
    def __init__(self,file_name,sheet_name):
        self.wb = load_workbook(file_name)
        self.sh=self.wb[sheet_name]

    def header(self):
        title = []
        for item in list(self.sh.rows)[0]:
            title.append(item.value)
        return title

    def read_all_datas(self):
        all_data=[]
        titles=self.header()
        for item in list(self.sh.rows)[1:]:  # 遍历数据行
            value = []
            for val in item: #获取每一行的数据
                value.append(val.value)
            res = dict(zip(titles,value))#title 和每一行数据打包成为字典
            # print(type(res))
            all_data.append(res) #追加数据
        return all_data

    @staticmethod
    def write_back(file_name, sheet_name, row, col, result):  # 专门写会数据
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]

        sheet.cell(row, col).value = result
        wb.save(file_name)  # 保存结果
        wb.close()

    def close_file(self):
        self.wb.close()

if __name__ == '__main__':
    do=HandlerExcel(conf.excel,"add")
    res=do.read_all_datas()
    for item in res:
        print(item)




